import datetime
import os
import time

from flask import Blueprint, jsonify, request, render_template, current_app
from sqlalchemy import func
from sqlalchemy import or_

from models.applicationsModel import db, Applications
from models.apptypesModel import AppTypes
from models.carsModel import Cars
from models.unitsModel import Units
from models.uursModel import Uurs
from models.signsModel import Signs

from models.maintainDataModel import MaintainData
from models.rechargeDataModel import RechargeData
from models.refuelDataModel import RefuelData
from models.repairDataModel import RepairData
from models.mtPhotosModel import MtPhotos

from models.flowsModel import Flows
from models.unitsAppTypesModel import UnitsApptypes
from models.rpPhotosModel import RpPhotos
from models.rfPhotosModel import RfPhotos

from schema.modelSchema import applications_schema_single, applications_schema, \
    app_types_schema_single, cars_schema_single, flows_schema, flows_schema_single, \
    recharge_data_schema_single, refuel_data_schema_single, maintain_data_schema_single, \
    repair_data_schema_single, maintain_photos_schema, repair_photos_schema, refuel_photos_schema, units_schema_single, \
    users_schema_single, signs_schema_single, maintain_data_schema, recharge_data_schema, \
    refuel_data_schema, repair_data_schema, cars_schema

from fsm.flow import Flow

from auth.identityAuth import role_auth_require, field_check_unit
from auth.msgDef import TYPE_NAME_TYPE_ID_REFLECTION, TYPE_NAME_TYPE_MEMO_REFLECTION, \
    TYPE_ID_TYPE_NAME_REFLECTION, FIELD_CODE

from utils.get_model import get_model
from utils.tool import operation_res_build, get_current_time, auth_role_exclude, file_name_picker, long_word_split
from utils.log import log_info_record, LOG_ACTIONS

import logger
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment

applications_bp = Blueprint("applications", __name__, url_prefix="/api/applications")


def get_data(app_id, type):
    """
    通过申请类型 查出data 以及图片
    :param app_id:
    :param type:申请的类型
    :return:
    """
    app_data = {}
    # photos_list = []
    if type == "maintaindata":
        mt_data = MaintainData.query.filter(app_id == MaintainData.app_id).first()
        if not mt_data is None:
            mt_data_ = maintain_data_schema_single.dump(mt_data)
            app_data.update(mt_data_)
            # 图片
            pre_photos = MtPhotos.query.filter(MtPhotos.mt_id == mt_data_["mt_id"]).all()
            if not len(pre_photos):
                return app_data
            # else:
            #     for photo in pre_photos:
            #         photos_list.append(maintain_photos_schema_single.dump(photo))
            # del
            # app_data["photos"]=photos_list

            app_data["photos"] = maintain_photos_schema.dump(pre_photos)
        return app_data

    elif type == "rechargedata":
        pre_data = RechargeData.query.filter(RechargeData.app_id == app_id).first()
        if pre_data is not None:
            app_data.update(recharge_data_schema_single.dump(pre_data))
        return app_data

    elif type == "refueldata":
        pre_data = RefuelData.query.filter(RefuelData.app_id == app_id).first()
        if pre_data is not None:
            data_ = refuel_data_schema_single.dump(pre_data)
            app_data.update(data_)
            # 查photo
            pre_photos = RfPhotos.query.filter(RfPhotos.ref_id == data_["ref_id"]).all()
            if not len(pre_photos):
                return app_data
            # else:
            #     for photo in pre_photos:
            #         photos_list.append(refuel_photos_schema.dump(photo))
            # app_data["photos"] = photos_list
            app_data["photos"] = refuel_photos_schema.dump(pre_photos)
        return app_data

    elif type == "repairdata":
        pre_data = RepairData.query.filter(RepairData.app_id == app_id).first()
        if pre_data is None:
            return app_data
        else:
            data_ = repair_data_schema_single.dump(pre_data)
            app_data.update(data_)
            pre_photos = RpPhotos.query.filter(RpPhotos.rp_id == data_["rp_id"]).all()
            if not len(pre_photos):
                return app_data
            # else:
            #     for photo in pre_photos:
            #         photos_list.append(refuel_photos_schema.dump(photo))
            # app_data["photos"] = photos_list
            app_data["photos"] = repair_photos_schema.dump(pre_photos)

        return app_data


@applications_bp.route("/app_id/<int:app_id>", methods=["GET"])
@role_auth_require(grant_all=True)
def app_request_by_id(role, identity, app_id):
    """
    app信息获取，根据app_id
    """
    tmp = {}
    try:
        if role == "system":
            app = Applications.query.filter(Applications.app_id == app_id).first()
        else:
            app = Applications.query.filter(Applications.app_id == app_id,
                                            Applications.uur_id == identity["uur_id"] if role == "employee" else True,
                                            Applications.unit_id_copy == identity[
                                                "unit_id"] if role != "system" else True).first()
        if app is None:
            return jsonify(operation_res_build("load app is not exist", False))

        app_ = applications_schema_single.dump(app)

        # 获取app最新的流程
        last_flow = Flows.query.filter(Flows.app_id == app_id).order_by(Flows.flowstate_id.desc()).first()
        _last_flow = None
        if last_flow is not None:
            _last_flow = flows_schema_single.dump(last_flow)
        else:
            # 没有初始的流程
            return jsonify(operation_res_build(f"this app {app_id} has no initial flow", False))

        car_ = cars_schema_single.dump(app.car)
        app_types_ = app_types_schema_single.dump(app.type)

        tmp[app_["type_name_copy"]] = get_data(app_["app_id"], app_["type_name_copy"])
        tmp.update(app_)
        tmp.update(car_)
        tmp["type_memo"] = app_types_["type_memo"]
        tmp["flowstate_title"] = _last_flow["flowstate_title"]
        tmp["flowstate_event"] = _last_flow["flowstate_event"]

        flow_fsm = Flow()
        tmp["flowstate"] = flow_fsm.STATES_SET[_last_flow["flowstate_title"]]
        tmp["_events"] = flow_fsm.get_events_by_status(tmp["flowstate"])

        tmp["flows"] = flows_schema.dump(
            Flows.query.filter(Flows.app_id == app_id).order_by(Flows.flowstate_id.asc()).all())
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "load app failed")
        logger.exception(e)
        return jsonify(operation_res_build("load app failed", False))

    return jsonify(operation_res_build("load app ok", True, data=tmp))


@applications_bp.route("/", methods=["POST"])
@role_auth_require(grant_all=True)
@log_info_record(LOG_ACTIONS["ADD"])
def app_add(role, identity):
    """
    添加app 并添加类型申请表中的数据
    :param role:
    :param identity:
    :return:
    """

    data = request.get_json()
    type_id = data.get("type_id")
    car_id = data.get("car_id")

    try:
        ua = UnitsApptypes.query.filter(UnitsApptypes.type_id == type_id,
                                        UnitsApptypes.unit_id == identity["unit_id"]).first()
        type_ = app_types_schema_single.dump(ua.type)

        unit_ = units_schema_single.dump(Units.query.filter(Units.unit_id == identity["unit_id"]).first())

        # 判断是否过期
        s1_1 = time.strptime(unit_["rent_expiretime"], "%Y-%m-%dT%H:%M:%S")
        s2_1 = time.strptime(get_current_time(), "%Y-%m-%d %H:%M:%S")

        if s1_1 < s2_1:
            return jsonify(operation_res_build("公司单位已超过租期时间，无法进行申请操作", False)), False

        # 公司是否是禁用状态
        if not unit_["unit_state"]:
            return jsonify(operation_res_build("公司单位处于禁用状态，无法进行申请操作", False)), False
    except AttributeError:
        return jsonify(operation_res_build(f"in uur_id: {identity['uur_id']} identity,"
                                           f" type_id: {type_id} does not exist", False)), False

    type_name_copy = type_["type_name"]
    app_sheet = None
    app_time_start = get_current_time()
    app_time_end = None
    app_quick = data.get("app_quick")
    flowstate_title_copy = ''
    flowstate_time_copy = None
    ur_id = None
    app_memo = data.get("app_memo")

    try:
        # 同一个车牌
        app = Applications.query.filter(Applications.car_id == car_id,
                                        Applications.type_name_copy == type_["type_name"],
                                        Applications.app_time_end == None).first()

        if app:
            return jsonify(operation_res_build(f"该车辆的{type_['type_memo']}已在进行中，请勿重复提交申请", False)), False

        app = Applications(identity['uur_id'], type_id, car_id, identity["name"], type_name_copy, app_quick,
                           identity['unit_id'], app_sheet, app_time_start, app_time_end, flowstate_title_copy,
                           flowstate_time_copy, ur_id, app_memo)

        db.session.add(app)
        db.session.flush()
        db.session.commit()

    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in addition")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in addition", False)), False

    return jsonify(operation_res_build("addition ok", True, app_id=app.app_id)), True


@applications_bp.route("/<int:app_id>", methods=["DELETE"])
@role_auth_require(auth_roles=["system", "unit_manager"])
@log_info_record(LOG_ACTIONS["DEL"])
def app_delete(role, identity, app_id):
    """
    app删除
    """
    try:
        app = Applications.query.filter(Applications.app_id == app_id,
                                        Applications.unit_id_copy == identity[
                                            "unit_id"] if role == "unit_manager" else True).first()

        if app is None:
            raise Exception

        db.session.delete(app)
        db.session.commit()
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in deletion")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in deletion", False)), False

    return jsonify(operation_res_build("deletion ok", True)), True


# @applications_bp.route("/todo", methods=["GET"])
# @role_auth_require(auth_roles=["unit_leader", "unit_officer", "subunit_officer"])
# def app_request_todo(role, identity):
#     try:
#         apps = Applications.query.filter(Applications.unit_id_copy == identity["unit_id"],
#                                          Applications.uur_id != identity["uur_id"],
#                                          Applications.app_time_end == None).all()  # 这里的None比较必须得用 ==，不能用 is
#
#         rets = []
#
#         for _ in apps:
#             app_ = applications_schema_single.dump(_)
#
#             is_add_flag = False
#
#             if role == "unit_leader":
#                 if app_["flowstate_title_copy"] == "LEADER_REVIEW" or app_["flowstate_title_copy"] == "INFO_CHECK":
#                     is_add_flag = True
#             else:
#                 if app_["flowstate_title_copy"] == "OFFICER_REVIEW":
#                     is_add_flag = True
#
#             if is_add_flag:
#                 ret = {}
#
#                 car_ = cars_schema_single.dump(_.car)
#                 type_ = app_types_schema_single.dump(_.type)
#
#                 ret.update(app_)
#                 # 添加data以及照片
#                 ret[app_["type_name_copy"]] = get_data(app_["app_id"], app_["type_name_copy"])
#
#                 ret.update(car_)
#                 ret["type_memo"] = type_["type_memo"]
#
#                 rets.append(ret)
#     except Exception as e:
#         logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
#         logger.exception(e)
#         return jsonify(operation_res_build("a error happened in operation", False))
#
#     return jsonify(operation_res_build("load apps ok", True, data=rets))


@applications_bp.route("/handling", methods=["GET"])
@role_auth_require(grant_all=True)
def app_request_handling(role, identity):
    try:
        apps = Applications.query.filter(Applications.unit_id_copy == identity["unit_id"],
                                         Applications.uur_id == identity["uur_id"],
                                         Applications.app_time_end == None).all()

        rets = []

        for _ in apps:
            app_ = applications_schema_single.dump(_)

            ret = {}
            car_ = cars_schema_single.dump(_.car)

            ret["app_id"] = app_["app_id"]
            ret["name_copy"] = app_["name_copy"]
            ret["app_time_start"] = app_["app_time_start"]
            ret["flowstate_title_copy"] = app_["flowstate_title_copy"]
            ret["palte_number"] = car_["palte_number"]
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]

            rets.append(ret)
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data=rets))


@applications_bp.route("/<int:app_id>", methods=["PUT"])
@role_auth_require(grant_all=True)
@log_info_record(LOG_ACTIONS["MOD"])
def app_update(role, identity, app_id):
    """
    app更新
    """

    data = request.get_json()

    # type_id = data.get("type_id")
    # car_id = data.get("car_id")
    # name_copy = data.get("name_copy")
    app_memo = data.get("app_memo")

    try:
        if role == "system":
            pre_sql = Applications.query.filter(Applications.app_id == app_id)
        else:
            pre_sql = Applications.query.filter(Applications.app_id == app_id,
                                                Applications.unit_id_copy == identity["unit_id"])

        app = pre_sql.first()

        if app is None:
            return jsonify(operation_res_build("the application does not exist", False)), False

        last_flow = Flows.query.filter(Flows.app_id == app_id).order_by(Flows.flowstate_id.desc()).first()
        if last_flow is None:
            return jsonify(operation_res_build("application's flows do not exist", False)), False
        else:
            _last_flow = flows_schema_single.dump(last_flow)

        # 流程状态机初始化
        flow_fsm = Flow()

        cur_state = flow_fsm.STATES_SET[_last_flow["flowstate_title"]]

        if cur_state != flow_fsm.STATES_SET["REJECT_APP"]:
            return jsonify(
                operation_res_build("application modification denied, current state no matchable", False)), False

        pre_sql.update({"app_memo": app_memo})

        db.session.commit()
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in update")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in update", False)), False

    return jsonify(operation_res_build("update ok", True)), True


@applications_bp.route("/done", methods=["GET"])
@role_auth_require(grant_all=True)
def app_request_done(role, identity):
    offset = request.args.get("offset")
    limit = request.args.get("limit")

    if offset is None or limit is None:
        return jsonify(operation_res_build("param miss", False))

    try:
        apps = Applications.query.order_by(Applications.app_id.desc()) \
            .filter(Applications.unit_id_copy == identity["unit_id"],
                    Applications.uur_id == identity["uur_id"],
                    Applications.app_time_end != None).offset(offset).limit(limit).all()

        total = db.session.query(func.count(Applications.app_id)) \
            .filter(Applications.unit_id_copy == identity["unit_id"],
                    Applications.uur_id == identity["uur_id"],
                    Applications.app_time_end != None).scalar()

        rets = []

        for _ in apps:
            app_ = applications_schema_single.dump(_)

            ret = {}
            car_ = cars_schema_single.dump(_.car)

            ret["app_id"] = app_["app_id"]
            ret["name_copy"] = app_["name_copy"]
            ret["app_time_end"] = app_["app_time_end"]
            ret["flowstate_title_copy"] = app_["flowstate_title_copy"]
            ret["palte_number"] = car_["palte_number"]
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]

            rets.append(ret)
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data={"rets": rets, "total": total}))


@applications_bp.route("/check/<int:app_id>", methods=["GET"])
@role_auth_require(auth_roles=auth_role_exclude("employee"))
def app_access_valid_check(role, identity, app_id):
    detail_type = request.args.get("detail_type")
    is_summary = request.args.get("is_summary")

    if detail_type is None:
        return jsonify(operation_res_build("params miss", False))
    elif int(detail_type) < 1 or int(detail_type) > 4:
        return jsonify(operation_res_build("params invalid", False))

    try:
        app = Applications.query.filter(Applications.app_id == app_id).first()

        if app is None:
            return jsonify(operation_res_build(f"app_id {app_id} doest not exist", False))

        app_ = applications_schema_single.dump(app)
        flows = Flows.query.filter(Flows.app_id == app_id).order_by(Flows.flowstate_id.asc()).all()
        flows_ = flows_schema.dump(flows)

        valid_flag = False

        if is_summary == '1':
            # 统计查看
            if role == "system":
                valid_flag = True
            else:
                # 垂直权限检测
                if field_check_unit(role, app_["unit_id_copy"], identity["unit_id"]) == FIELD_CODE["PASS"]:
                    valid_flag = True
        else:
            # 审核批准流程进入
            if role == "system" or ((app_["flowstate_title_copy"] == "OFFICER_REVIEW") and (detail_type == '1') and (
                    role == "subunit_officer" or role == "unit_officer")) or \
                    ((app_["flowstate_title_copy"] == "LEADER_REVIEW") and (detail_type == '2') and (
                            role == "unit_leader")) or \
                    ((app_["flowstate_title_copy"] == "INFO_CHECK") and (detail_type == '3') and (
                            role == "subunit_officer" or role == "unit_officer")):
                valid_flag = True

        if not valid_flag:
            return jsonify(operation_res_build("access is invalid", False))

        ret = {}
        app_ = applications_schema_single.dump(app)
        car_ = cars_schema_single.dump(app.car)
        type_ = app_types_schema_single.dump(app.type)

        extra_info = None
        if type_["type_name"] == "maintaindata":
            tmp = app.maintaindata

            if len(tmp):
                extra_info = maintain_data_schema_single.dump(tmp[0])
                extra_info["photos"] = maintain_photos_schema.dump(
                    MtPhotos.query.filter(MtPhotos.mt_id == extra_info["mt_id"]).all())
        elif type_["type_name"] == "rechargedata":
            tmp = app.rechargedata

            if len(tmp):
                extra_info = recharge_data_schema_single.dump(app.rechargedata[0])
                extra_info["photos"] = []
        elif type_["type_name"] == "refueldata":
            tmp = app.refueldata

            if len(tmp):
                extra_info = refuel_data_schema_single.dump(app.refueldata[0])
                extra_info["photos"] = refuel_photos_schema.dump(
                    RfPhotos.query.filter(RfPhotos.ref_id == extra_info["ref_id"]).all())
        elif type_["type_name"] == "repairdata":
            tmp = app.repairdata

            if len(tmp):
                extra_info = repair_data_schema_single.dump(app.repairdata[0])
                extra_info["photos"] = repair_photos_schema.dump(
                    RpPhotos.query.filter(RpPhotos.rp_id == extra_info["rp_id"]).all())

        ret.update(app_)
        ret.update(car_)
        ret["flows"] = flows_
        ret["type_memo"] = type_["type_memo"]
        ret["extraInfos"] = extra_info
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("access is valid", True, data=ret))


@applications_bp.route("/pending_application", methods=["GET"])
@role_auth_require(auth_roles=["subunit_officer", "unit_officer", "unit_manager", "system"])
def app_request_pending_application(role, identity):
    query = request.args.get("query")
    unit_id = request.args.get("unit_id")

    if role == "system" and unit_id is None:
        return jsonify(operation_res_build("unit_id is required", False))

    try:
        unit_id = int(unit_id)
    except ValueError:
        return jsonify(operation_res_build("error request params", False))

    try:
        if query is None:
            apps = Applications.query.filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "OFFICER_REVIEW").all()
        else:
            apps = Applications.query.join(AppTypes, Cars) \
                .filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "OFFICER_REVIEW",
                or_(
                    True if query is None else AppTypes.type_memo.ilike(f"%{query}%"),
                    True if query is None else Applications.name_copy.ilike(f"%{query}%"),
                    True if query is None else or_(Cars.car_brand.ilike(f"%{query}%"),
                                                   Cars.palte_number.ilike(f"%{query}%"),
                                                   Cars.car_type.ilike(f"%{query}%")),
                    True if query is None else Applications.app_time_start.ilike(f"%{query}%")
                )).all()

        rets = []

        for _ in apps:
            ret = {}

            app_ = applications_schema_single.dump(_)
            car_ = cars_schema_single.dump(_.car)

            ret.update(app_)
            ret.update(car_)
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]

            rets.append(ret)
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data=rets))


@applications_bp.route("/pending_approval", methods=["GET"])
@role_auth_require(auth_roles=["unit_leader", "unit_manager", "system"])
def app_request_pending_approval(role, identity):
    query = request.args.get("query")
    unit_id = request.args.get("unit_id")

    if role == "system" and unit_id is None:
        return jsonify(operation_res_build("unit_id is required", False))

    try:
        if query is None:
            apps = Applications.query.filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "LEADER_REVIEW").all()
        else:
            apps = Applications.query.join(AppTypes, Cars) \
                .filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "LEADER_REVIEW",
                or_(
                    True if query is None else AppTypes.type_memo.ilike(f"%{query}%"),
                    True if query is None else Applications.name_copy.ilike(f"%{query}%"),
                    True if query is None else or_(Cars.car_brand.ilike(f"%{query}%"),
                                                   Cars.palte_number.ilike(f"%{query}%"),
                                                   Cars.car_type.ilike(f"%{query}%")),
                    True if query is None else Applications.app_time_start.ilike(f"%{query}%")
                )).all()

        rets = []

        for _ in apps:
            ret = {}

            app_ = applications_schema_single.dump(_)
            car_ = cars_schema_single.dump(_.car)

            ret.update(app_)
            ret.update(car_)
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]

            rets.append(ret)
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data=rets))


@applications_bp.route("/pending_audit", methods=["GET"])
@role_auth_require(auth_roles=["subunit_officer", "unit_officer", "unit_manager", "system"])
def app_request_pending_audit(role, identity):
    query = request.args.get("query")
    unit_id = request.args.get("unit_id")

    if role == "system" and unit_id is None:
        return jsonify(operation_res_build("unit_id is required", False))

    try:
        if query is None:
            apps = Applications.query.filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "INFO_CHECK").all()
        else:
            apps = Applications.query.join(AppTypes, Cars) \
                .filter(
                Applications.ur_id == identity["ur_id"] if role != "system" else Applications.unit_id_copy == unit_id,
                Applications.app_time_end == None,
                Applications.flowstate_title_copy == "INFO_CHECK",
                or_(
                    True if query is None else AppTypes.type_memo.ilike(f"%{query}%"),
                    True if query is None else Applications.name_copy.ilike(f"%{query}%"),
                    True if query is None else or_(Cars.car_brand.ilike(f"%{query}%"),
                                                   Cars.palte_number.ilike(f"%{query}%"),
                                                   Cars.car_type.ilike(f"%{query}%")),
                    True if query is None else Applications.app_time_start.ilike(f"%{query}%")
                )).all()

        rets = []

        for _ in apps:
            ret = {}

            app_ = applications_schema_single.dump(_)
            car_ = cars_schema_single.dump(_.car)

            ret.update(app_)
            ret.update(car_)
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]

            rets.append(ret)
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data=rets))


@applications_bp.route("/app_sum/<int:unit_id>", methods=["GET"])
@role_auth_require(auth_roles=auth_role_exclude("employee"))
def app_sum_request(role, identity, unit_id):
    query = request.args.get("query")
    type_id = int(request.args.get("type_id"))

    offset = request.args.get("offset")
    limit = request.args.get("limit")

    if offset is None or limit is None:
        return jsonify(operation_res_build("request params miss", False))

    if role != "system" and identity["unit_id"] != unit_id:
        # 垂直权限检测
        if field_check_unit(role, unit_id, identity["unit_id"]) != FIELD_CODE["PASS"]:
            return jsonify(operation_res_build("no right to access this resources", False))

    try:
        if query is None:
            total = db.session.query(func.count(Applications.app_id)) \
                .filter(Applications.unit_id_copy == unit_id,
                        Applications.type_id == type_id if type_id else True).scalar()

            apps = Applications.query.filter(Applications.unit_id_copy == unit_id,
                                             Applications.type_id == type_id if type_id else True) \
                .order_by(Applications.app_id.desc()).offset(offset).limit(limit).all()
        else:
            total = db.session.query(func.count(Applications.app_id)).join(AppTypes, Cars) \
                .filter((Applications.unit_id_copy == unit_id) &
                        (Applications.type_id == type_id if type_id else True) &
                        (
                                AppTypes.type_memo.ilike(f"%{query}%") |
                                Applications.name_copy.ilike(f"%{query}%") |
                                Cars.car_brand.ilike(f"%{query}%") |
                                Cars.palte_number.ilike(f"%{query}%") |
                                Cars.car_type.ilike(f"%{query}%") |
                                Applications.app_time_start.ilike(f"%{query}%")
                        )).scalar()

            apps = Applications.query.join(AppTypes, Cars) \
                .filter(Applications.unit_id_copy == unit_id,
                        Applications.type_id == type_id if type_id else True,
                        or_(
                            True if query is None else AppTypes.type_memo.ilike(f"%{query}%"),
                            True if query is None else Applications.name_copy.ilike(f"%{query}%"),
                            True if query is None else or_(Cars.car_brand.ilike(f"%{query}%"),
                                                           Cars.palte_number.ilike(f"%{query}%"),
                                                           Cars.car_type.ilike(f"%{query}%")),
                            True if query is None else Applications.app_time_start.ilike(f"%{query}%")
                        )) \
                .order_by(Applications.app_id.desc()).offset(offset).limit(limit).all()

        rets = []

        for _ in apps:
            ret = {}

            app_ = applications_schema_single.dump(_)
            car_ = cars_schema_single.dump(_.car)

            ret.update(app_)
            ret.update(car_)
            ret["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]]
            ret["type_id"] = TYPE_NAME_TYPE_ID_REFLECTION[app_["type_name_copy"]]
            rets.append(ret)

    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("load apps ok", True, data=rets, total=total))


@applications_bp.route("/app_print/<int:app_id>", methods=["GET"])
@role_auth_require(grant_all=True)
def app_print(role, identity, app_id):
    try:
        pre_sql = Applications.query.filter(Applications.app_id == app_id)
        app = pre_sql.first()

        if app is None:
            return jsonify(operation_res_build(f"app {app_id} does not exist", False))

        app_ = applications_schema_single.dump(app)
        if role != "system" and field_check_unit(role, app_["unit_id_copy"], identity["unit_id"]) != FIELD_CODE["PASS"]:
            return jsonify(operation_res_build("no right to access this resources", False))
        elif not app_["app_time_end"]:
            return jsonify(operation_res_build(f"app_id {app_['app_id']} 还未完成", False))

        create_flag = True

        if app_["app_sheet"] is not None and app_["app_sheet"].strip() != '':
            if os.path.exists(
                    os.path.join(os.path.join(current_app.config["PRINT_STORAGE_BASE_PATH"],
                                              str(app_["unit_id_copy"]) + '/',
                                              str(app_["app_sheet"])))
            ):
                create_flag = False

        if create_flag:
            print_file_name = file_name_picker(app_["uur_id"], app_["app_id"], ".pdf")
            print_file_path = os.path.join(current_app.config["PRINT_STORAGE_BASE_PATH"],
                                           str(app_["unit_id_copy"]) + '/')

            # 创建文件夹
            if not os.path.exists(print_file_path):
                os.makedirs(print_file_path)

            print_file_path = os.path.join(print_file_path, print_file_name)
            print_file = open(print_file_path, "wb")

            unit = Units.query.filter(Units.unit_id == app_["unit_id_copy"]).first()
            if unit is None:
                return jsonify(operation_res_build(f"unit {app_['unit_id_copy']} does not exist", False))

            unit_ = units_schema_single.dump(unit)
            car_ = cars_schema_single.dump(app.car)

            # 打印图片获取
            if not (pics_path_set := __do_print_pics_fetch(app)):
                return jsonify(operation_res_build(f"app_id {app_['app_id']} has not finished yet", False))

            flows_ = flows_schema.dump(
                Flows.query.filter(Flows.app_id == app_id).order_by(Flows.flowstate_id.desc()).all())

            # 获取关键人物信息
            driver = None
            officer = {"name": '', "comment": '', "sign_img": '', "sign_time": ''}
            leader = {"name": '', "comment": '', "sign_img": '', "sign_time": ''}

            for flow_ in flows_:
                if flow_["flowstate_event"] == "SUBMIT" or flow_["flowstate_event"] == "EMERGENCY_APP_FLY":
                    driver = __do_user_info_fetch(flow_["uur_id"], flow_)
                elif flow_["flowstate_event"] == "OFFICER_APPROVAL":
                    officer = __do_user_info_fetch(flow_["uur_id"], flow_)
                elif flow_["flowstate_event"] == "LEADER_APPROVAL":
                    leader = __do_user_info_fetch(flow_["uur_id"], flow_)

            if not driver:
                return jsonify(operation_res_build(f"app_id {app_['app_id']} miss a significant flow", False))

            # PDF生成
            pisa.CreatePDF(
                render_template("index.html", unit_name_ls=long_word_split(unit_["unit_name"], 6),
                                type_memo=TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]],
                                type_memo_=TYPE_NAME_TYPE_MEMO_REFLECTION[app_["type_name_copy"]][: 2],
                                car_type=car_["car_type"], name_copy=app_["name_copy"],
                                palte_number=car_["palte_number"], app_memo_ls=long_word_split(app_["app_memo"], 30),
                                officer=officer,
                                driver=driver,
                                leader=leader,
                                pics=pics_path_set),

                dest=print_file
            )

            print_file.close()

            pre_sql.update({
                "app_sheet": print_file_name
            })
            db.session.commit()
        else:
            print_file_path = os.path.join(current_app.config["PRINT_STORAGE_BASE_PATH"],
                                           str(app_["unit_id_copy"]) + '/',
                                           str(app_["app_sheet"]))
    except Exception as e:
        print(e)
        return jsonify(operation_res_build("a error happened in operation", False))

    return jsonify(operation_res_build("print file generated ok", True, path=print_file_path))


def __do_user_info_fetch(uur_id, flow_):
    tmp = Uurs.query.filter(Uurs.uur_id == uur_id).first()
    if tmp is None:
        return jsonify(operation_res_build(f"unknown uur", False))

    user_ = users_schema_single.dump(tmp.u)

    sign = Signs.query.filter(Signs.u_id == user_["u_id"]).first()
    sign_img = None

    if sign:
        sign_img = signs_schema_single.dump(sign)["sign_template"]

    sign_time = flow_["flowstate_time"].replace('T', ' ').split(' ')[0]
    sign_time_arr = sign_time.split('-')
    return {"name": user_["name"], "comment": flow_["comment"],
            "sign_img": sign_img, "sign_time": f"{sign_time_arr[0]}年{sign_time_arr[1]}月{sign_time_arr[2]}日"}


def __do_print_pics_fetch(app):
    # 所有图片地址获取
    specify_data = []
    specify_data_schema = None
    specify_attach_schema_many = None
    col_name_data = ""
    col_name_attach = ""
    attachments_pics = []
    if not (specify_data := app.maintaindata):
        if not (specify_data := app.refueldata):
            if not (specify_data := app.rechargedata):
                if not (specify_data := app.repairdata):
                    return None
                else:
                    specify_data_schema = repair_data_schema_single
                    specify_attach_schema_many = repair_photos_schema
                    col_name_data = "photo_rp_receipt"

                    attachments_pics = specify_data[0].rpphotos
                    col_name_attach = "prp_file"
            else:
                specify_data_schema = recharge_data_schema_single
                col_name_data = "photo_rc_receipt"
        else:
            specify_data_schema = refuel_data_schema_single
            specify_attach_schema_many = refuel_photos_schema
            col_name_data = "photo_ref_receipt"

            attachments_pics = specify_data[0].rfphotos
            col_name_attach = "prf_file"
    else:
        specify_data_schema = maintain_data_schema_single
        specify_attach_schema_many = maintain_photos_schema
        col_name_data = "photo_mt_receipt"

        attachments_pics = specify_data[0].mtphotos
        col_name_attach = "pmt_file"

    specify_data_ = specify_data_schema.dump(specify_data[0])
    receipt_pic_path = specify_data_[col_name_data]

    attachments_pics_path = []
    if col_name_data != "photo_rc_receipt":
        attachments_pics_path = [x[col_name_attach] for x in specify_attach_schema_many.dump(attachments_pics)]

    return [receipt_pic_path] + attachments_pics_path


@applications_bp.route("/types_sum/<int:unit_id>", methods=["GET"])
@role_auth_require(grant_all=True)
def types_sum(role, identity, unit_id):
    type_id = int(request.args.get("type_id"))
    start_time = request.args.get("start_time")
    end_time = request.args.get("end_time")
    car_id = request.args.get("car_id")
    is_export = int(request.args.get("is_export")) if request.args.get("is_export") else 0

    if role != "system" and field_check_unit(role, unit_id, identity["unit_id"]) != FIELD_CODE["PASS"]:
        return jsonify(operation_res_build("no right to access this resources", False))

    if TYPE_ID_TYPE_NAME_REFLECTION[type_id] == "refueldata":
        ref_self = request.args.get("ref_self")

        if ref_self is None:
            return jsonify(operation_res_build("request param miss", False))

        ref_self = int(ref_self)

    try:
        if start_time:
            s_time = time.strptime(start_time, "%Y-%m-%d")
        if end_time:
            # %H:%M:%S
            e_time = time.strptime(end_time, "%Y-%m-%d %H:%M:%S")

        type_name = TYPE_ID_TYPE_NAME_REFLECTION[type_id]
    except Exception:
        return jsonify(operation_res_build("error params", False))

    if is_export:
        apps = Applications.query.filter(Applications.unit_id_copy == unit_id, Applications.type_id == type_id,
                                         Applications.app_time_end != None, Applications.flowstate_title_copy != "WITHDRAW",
                                         Applications.app_time_end >= s_time if start_time else True,
                                         Applications.app_time_end <= e_time if end_time else True).all()
    else:
        apps = Applications.query.filter(Applications.unit_id_copy == unit_id, Applications.type_id == type_id,
                                         Applications.app_time_end != None,
                                         Applications.flowstate_title_copy != "WITHDRAW",
                                         Applications.app_time_end >= s_time if start_time else True,
                                         Applications.app_time_end <= e_time if end_time else True) \
            .offset(request.args.get("offset")).limit(request.args.get("limit")).all()

        total = db.session.query(func.count(Applications.app_id)).filter(
            Applications.unit_id_copy == unit_id, Applications.type_id == type_id,
            Applications.app_time_end != None, Applications.flowstate_title_copy != "WITHDRAW",
            Applications.app_time_end >= s_time if start_time else True,
            Applications.app_time_end <= e_time if end_time else True
        ).scalar()

    apps_ = applications_schema.dump(apps)

    if car_id:
        apps_ = [x for x in apps_ if x["car_id"] == int(car_id)]

    apps_ids = [x["app_id"] for x in apps_]

    # 加油统计需要过滤是否本车
    if TYPE_ID_TYPE_NAME_REFLECTION[type_id] == "refueldata":
        refs_ = refuel_data_schema.dump(
            RefuelData.query.filter(
                RefuelData.app_id.in_(apps_ids),
                RefuelData.ref_self == ref_self if ref_self != -1 else True
            ).all()
        )

        refs_ids = [ref_["app_id"] for ref_ in refs_]
        apps_ = [app_ for app_ in apps_ if (app_["app_id"] in refs_ids)]

    specify_data = None
    specify_data_schema = None
    to_del_col = ''
    if type_name == "maintaindata":
        specify_data = MaintainData
        specify_data_schema = maintain_data_schema
        to_del_col = "photo_mt_receipt"
    elif type_name == "rechargedata":
        specify_data = RechargeData
        specify_data_schema = recharge_data_schema
        to_del_col = "photo_rc_receipt"
    elif type_name == "refueldata":
        specify_data = RefuelData
        specify_data_schema = refuel_data_schema
        to_del_col = "photo_ref_receipt"
    elif type_name == "repairdata":
        specify_data = RepairData
        specify_data_schema = repair_data_schema
        to_del_col = "photo_rp_receipt"

    apps_ids = [x["app_id"] for x in apps_]
    cars_ids = set([x["car_id"] for x in apps_])
    specify_data_set = specify_data_schema.dump(specify_data.query.filter(specify_data.app_id.in_(apps_ids)).all())
    cars_tmp = cars_schema.dump(Cars.query.filter(Cars.car_id.in_(cars_ids)).all())
    cars_ = {}

    for car_ in cars_tmp:
        del car_["capital_balance"]
        cars_[car_["car_id"]] = car_

    if len(apps_) != len(specify_data_set):
        return jsonify(operation_res_build("abnormal data", False))

    apps_ = sorted(apps_, key=lambda x: x["app_id"])
    specify_data_set = sorted(specify_data_set, key=lambda x: x["app_id"])

    for i in range(len(apps_)):
        del apps_[i]["app_sheet"]
        del specify_data_set[i][to_del_col]

        apps_[i]["data"] = specify_data_set[i]
        apps_[i]["car"] = cars_[apps_[i]["car_id"]]

    if is_export:
        return __do_app_export(type_name, apps_, unit_id, car_id, start_time, end_time)
    else:
        return jsonify(operation_res_build("load ok", True, data=apps_, total=total))


def __do_app_export(type_name, apps_, unit_id, car_id, start_time, end_time):
    export_file_path = os.path.join(current_app.config["EXPORT_STORAGE_BASE_PATH"],
                                    f"{TYPE_NAME_TYPE_MEMO_REFLECTION[type_name]}_{time.time()}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers, tmp_col_names, count_col_name_cost, unit_name, car_palte_number, merge_right_index = \
        __do_export_headers_fetch(type_name, unit_id, car_id)

    # 表头写入
    ws.merge_cells(f"A1:{merge_right_index}1")
    type_ = TYPE_NAME_TYPE_MEMO_REFLECTION[type_name][: 2]
    ws["A1"] = unit_name + type_ + "统计表"
    ws["A1"].font = Font(u'宋体', size=18)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 16.25  # 135px
    ws.column_dimensions['D'].width = 20  # 165px
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    if type_name == "maintaindata":
        ws.column_dimensions['H'].width = 14.38  # 120px
        ws.column_dimensions['I'].width = 18.13  # 150px
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 11.88  # 100px
    elif type_name == "rechargedata":
        ws.column_dimensions['I'].width = 11.88  # 100px
        ws.column_dimensions['J'].width = 20
    elif type_name == "refueldata":
        ws.column_dimensions['H'].width = 11.25  # 95px
        ws.column_dimensions['I'].width = 14.38  # 120px
        ws.column_dimensions['J'].width = 14.38
        ws.column_dimensions['K'].width = 9.38   # 80px
        ws.column_dimensions['L'].width = 11.25  # 95px
        ws.column_dimensions['M'].width = 20  # 165px
        ws.column_dimensions['N'].width = 20  # 165px
    elif type_name == "repairdata":
        ws.column_dimensions['H'].width = 14.38  # 120px
        ws.column_dimensions['I'].width = 11.88  # 100px
        ws.column_dimensions['J'].width = 20  # 165px
        ws.column_dimensions['K'].width = 18.13  # 150px

    ws.cell(2, 1).value = "车牌号：" + ("所有" if not car_id else car_palte_number)

    tmp = "所有"
    if end_time:
        end_time = end_time.split(' ')[0]
    if start_time or end_time:
        if start_time != end_time:
            tmp = start_time + " 至 " + end_time
        else:
            tmp = start_time
    ws.cell(2, 2).value = f"日期范围：{tmp}"
    for i, header in enumerate(headers):
        ws.cell(3, 1 + i).value = header

    count_refs_volume_total = 0
    count_refs_costs_total = 0
    count_mt_costs_total = 0
    count_rc_costs_total = 0
    count_rp_costs_total = 0

    # 申请集时间格式转换
    for app_ in apps_:
        app_["app_time_start"] = str(datetime.datetime.strptime(app_["app_time_start"], "%Y-%m-%dT%H:%M:%S"))

        if app_["app_time_end"]:
            app_["app_time_end"] = str(datetime.datetime.strptime(app_["app_time_end"], "%Y-%m-%dT%H:%M:%S"))

        try:
            app_["data"]["mt_time"] = str(datetime.datetime.strptime(app_["data"]["mt_time"], "%Y-%m-%dT%H:%M:%S"))
        except KeyError:
            pass

        try:
            app_["data"]["rc_time"] = str(datetime.datetime.strptime(app_["data"]["rc_time"], "%Y-%m-%dT%H:%M:%S"))
        except KeyError:
            pass

        try:
            app_["data"]["ref_time"] = str(datetime.datetime.strptime(app_["data"]["ref_time"], "%Y-%m-%dT%H:%M:%S"))
        except KeyError:
            pass

        try:
            app_["data"]["rp_time"] = str(datetime.datetime.strptime(app_["data"]["rp_time"], "%Y-%m-%dT%H:%M:%S"))
        except KeyError:
            pass

    # 特殊处理
    if type_name == "refueldata":
        # RULE: (车辆ID, 车牌, 加油量, 加油公里, 加油花费, 时间, 是否为本车)
        vehicle_info = [(app_["car"]["car_id"], app_["car"]["palte_number"],
                         app_["data"]["ref_volume"], app_["data"]["ref_mile"], app_["data"]["ref_cost"],
                         app_["data"]["ref_time"], app_["data"]["ref_self"])
                        for app_ in apps_]

        vehicle_info, single_vehicle_info = list(__do_elem_in_vehicle(vehicle_info))
        # 按车编号和时间排序
        vehicle_info = sorted(vehicle_info,
                              key=lambda vehicle: (
                              vehicle[0], datetime.datetime.strptime(vehicle[5], "%Y-%m-%d %H:%M:%S")))
        single_vehicle_info = sorted(single_vehicle_info, key=lambda vehicle: (
            vehicle[0], datetime.datetime.strptime(vehicle[5], "%Y-%m-%d %H:%M:%S")))

    # 内容写入
    for i, app_ in enumerate(apps_):
        data_ls = []

        for x in tmp_col_names:
            try:
                data_ls.append(app_["data"][x])
            except KeyError:
                # 可能是在car属性中
                data_ls.append(app_["car"][x])

        if type_name == "refueldata":
            count_refs_volume_total += app_["data"]["ref_volume"]
            count_refs_costs_total += app_["data"]["ref_cost"]
        elif type_name == "maintaindata":
            count_mt_costs_total += app_["data"]["mt_cost"]
        elif type_name == "rechargedata":
            count_rc_costs_total += app_["data"]["rc_value"]
        else:
            count_rp_costs_total += app_["data"]["rp_cost"]

        ws.append([1 + i,
                   # 申请信息
                   app_["name_copy"],
                   TYPE_NAME_TYPE_MEMO_REFLECTION[type_name],
                   app_["app_time_start"],
                   app_["app_time_end"],
                   app_["app_memo"],
                   # 车辆信息
                   app_["car"]["palte_number"],
                   # DATA信息
                   ] + data_ls)

    ws.append([])
    if type_name == "maintaindata":
        row_ = [f"汇总:", None, None, None, None, None, None, None, None, None, count_mt_costs_total]
    elif type_name == "rechargedata":
        row_ = [f"汇总:", None, None, None, None, None, None, None, count_rc_costs_total]
    elif type_name == "refueldata":
        row_ = [f"汇总:", None, None, None, None, None, None, None, None, None, count_refs_volume_total,
                count_refs_costs_total]
    elif type_name == "repairdata":
        row_ = [f"汇总:", None, None, None, None, None, None, None, count_rp_costs_total]
    ws.append(row_)

    if type_name == "refueldata":
        # RULE: (车辆ID, 车牌, 加油量统计, 加油公里统计, 加油花费统计)
        summary_ls = []
        miles_tmp_ls = []
        refs_tmp_ls = {}
        left_mile_scale = 0
        right_mile_scale = 0
        last_ref_volume = 0

        for vehicle in vehicle_info:
            count_in_flag = False
            refs_tmp_ls[vehicle[0]] = vehicle[2]

            for s in summary_ls:
                if s[0] == vehicle[0]:
                    count_in_flag = True

                    s[2] += vehicle[2]  #
                    refs_tmp_ls[vehicle[0]] = vehicle[2]
                    right_mile_scale = vehicle[3]  # 不断推进右边界加油公里
                    s[4] += vehicle[4]
                    break

            if not count_in_flag:
                miles_tmp_ls.append(right_mile_scale - left_mile_scale)  # 上个车辆的公里统计

                left_mile_scale = vehicle[3]
                summary_ls.append([vehicle[0], vehicle[1], vehicle[2], 0, vehicle[4]])

        if len(summary_ls) or len(single_vehicle_info):
            ws.append([])
            ws.append(["各车辆本车加油及油耗概况:"])
            ws.append(["车牌号", "总加油量", "总公里数", "总花费", "平均油耗", "备注：平均油耗计算未加入最后一次加油量"])

        if len(summary_ls):
            del miles_tmp_ls[0]  # 删除第一位冗余
            miles_tmp_ls.append(right_mile_scale - left_mile_scale)  # 最后一位的手动加上

            for s, m in zip(summary_ls, miles_tmp_ls):
                s[3] = m

            for row in summary_ls:
                err_row = [row[1], f"{row[2]}升", f"{row[3]}公里", f"￥{row[4]}元", "错误"]
                try:
                    if row[2] / row[3] <= 0:
                        ws.append(err_row)
                    else:
                        ws.append([row[1], f"{round(row[2], 3)}升", f"{round(row[3], 3)}公里", f"￥{round(row[4], 3)}元",
                                   f"{round((row[2] - refs_tmp_ls[row[0]]) * 100 / row[3], 3)}升/百公里"])
                except ZeroDivisionError:
                    ws.append(err_row)

        if len(single_vehicle_info):
            for row in single_vehicle_info:
                ws.append([row[1], f"{row[2]}升", None, f"￥{row[4]}元", None])
    wb.save(export_file_path)

    # 过期文件清理
    cur_time = time.time()
    for file_name in os.listdir(current_app.config["EXPORT_STORAGE_BASE_PATH"]):
        if cur_time - float(file_name[file_name.find('_') + 1: file_name.rfind('.')]) > current_app.config[
            "EXPORTED_FILE_EXIST"]:
            if os.path.exists(
                    to_deleted_file := os.path.join(current_app.config["EXPORT_STORAGE_BASE_PATH"], file_name)):
                os.remove(to_deleted_file)

    return jsonify(operation_res_build("export ok", True, path=export_file_path))


def __do_export_headers_fetch(type_name, unit_id, car_id):
    headers = ["编号", "申请人", "申请类型", "申请开始时间", "申请结束时间", "申请备注"]
    cars_headers = ["车牌号"]

    tmp_headers = []
    tmp_col_names = []
    count_col_name_cost = ''
    merge_right_index = 'U'
    if type_name == "maintaindata":
        tmp_headers = ["保养内容", "保养地址", "保养时间", "保养花费(元)", "保养时里程表(公里)"]
        tmp_col_names = ["mt_item", "mt_addr", "mt_time", "mt_cost", "mt_mile"]
        count_col_name_cost = "mt_cost"
        merge_right_index = 'L'
    elif type_name == "rechargedata":
        tmp_headers = ["加油卡号", "充值金额(元)", "充值时间"]
        tmp_col_names = ["card_number", "rc_value", "rc_time"]
        count_col_name_cost = "rc_value"
        merge_right_index = 'J'
    elif type_name == "refueldata":
        tmp_headers = ["加油卡号", "是否为本车加油", "油品单价(元/升)", "加油量(升)", "加油花费(元)", "加油时间", "加油地址",
                       "油品标号", "加油时里程表(公里)"]
        tmp_col_names = ["card_number", "ref_self", "ref_price", "ref_volume", "ref_cost",
                         "ref_time", "ref_addr", "ref_gas_type", "ref_mile"]
        count_col_name_cost = "ref_cost"
        merge_right_index = 'P'
    elif type_name == "repairdata":
        tmp_headers = ["维修内容", "维修花费(元)", "维修时间", "维修地点", "维修时里程表(公里)"]
        tmp_col_names = ["rp_item", "rp_cost", "rp_time", "rp_addr", "rp_mile"]
        count_col_name_cost = "rp_cost"
        merge_right_index = 'L'

    headers = headers + cars_headers + tmp_headers

    unit_name = ""
    try:
        unit_name = units_schema_single.dump(Units.query.filter(Units.unit_id == unit_id).first())["unit_name"]
    except Exception:
        pass

    car_palte_number = ""
    try:
        if car_id:
            car_palte_number = cars_schema_single.dump(Cars.query.filter(Cars.car_id == car_id).first())["palte_number"]
    except Exception:
        pass

    return headers, tmp_col_names, count_col_name_cost, unit_name, car_palte_number, merge_right_index


def __do_elem_in_vehicle(vehicle_info):
    seen = set()
    dupes_ls = []

    for index, vehicle in enumerate(vehicle_info):
        flag = False

        if vehicle[0] in seen and vehicle[6]:
            dupes_ls.append(vehicle)
        else:
            try:
                for x in vehicle_info[index + 1:]:
                    if x[0] == vehicle[0] and x[6]:
                        flag = True
                        break

                if vehicle[6]:
                    seen.add(vehicle[0])
            except IndexError:
                pass

        if flag and vehicle[6]:
            dupes_ls.append(vehicle)

    _ = [x[0] for x in dupes_ls]
    single_ls = [vehicle for vehicle in vehicle_info if vehicle[0] not in _ and vehicle[6]]

    return dupes_ls, single_ls


@applications_bp.route("/todoList", methods=["GET"])
@role_auth_require(grant_all=True)
def get_todo_app(role, identity):
    car_need = int(request.args.get("car_need"))

    try:
        self_handled_states = ["INFO_UPLOAD", "FULFILL_INFO", "REJECT_APP"]
        app_self = Applications.query.filter(Applications.uur_id == identity["uur_id"],
                                             Applications.flowstate_title_copy.in_(self_handled_states),
                                             Applications.app_time_end == None)

        app_self_ = applications_schema.dump(app_self)

        for x in app_self_:
            x["self_app"] = 1
            x["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[x["type_name_copy"]]

        app_approval = Applications.query.filter(Applications.ur_id == identity["ur_id"])
        app_approval_ = applications_schema.dump(app_approval)

        for x in app_approval_:
            x["self_app"] = 0
            x["type_memo"] = TYPE_NAME_TYPE_MEMO_REFLECTION[x["type_name_copy"]]

        ret = app_self_ + app_approval_

        if car_need:
            car_ids = set([x["car_id"] for x in ret])

            cars = Cars.query.filter(Cars.car_id.in_(car_ids)).all()
            cars_ = cars_schema.dump(cars)

            cars_reflection = {}
            for x in cars_:
                cars_reflection[x["car_id"]] = x["palte_number"]

            for x in ret:
                x["palte_number"] = cars_reflection[x["car_id"]]
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False)), False

    return jsonify(operation_res_build("load ok", True, data=ret))


@applications_bp.route("/del_receipt", methods=["DELETE"])
@role_auth_require(grant_all=True)
@log_info_record(LOG_ACTIONS["DEL"])
def del_receipt(role, identity):
    """
{
        "app_id":37,
}
    """
    data = request.get_json()
    app_id = data.get("app_id")
    # photo_upload_session = session.get("photo_upload_session")
    #
    # if photo_upload_session["uur_id"] != identity["uur_id"]:
    #     return jsonify(operation_res_build("此申请不存在！", False)), False
    try:
        apps = Applications.query.filter(Applications.app_id == app_id, Applications.app_time_end == None, ).first()
        if apps is None:
            return jsonify(operation_res_build("此申请不存在！", False)), False
        app_ = applications_schema_single.dump(apps)

        if int(app_["uur_id"]) != int(identity["uur_id"]):
            return jsonify(operation_res_build("此申请不存在！", False)), False
        flag = delete_receipt(app_)
        if flag == True:
            return jsonify(operation_res_build("删除成功！", True)), True
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False)), False


@applications_bp.route("/del_photos", methods=["DELETE"])
@role_auth_require(grant_all=True)
@log_info_record(LOG_ACTIONS["DEL"])
def del_photos(role, identity):
    """
    {
        "app_id": 1,
        "photo_del_id":[1,3,4,5,6]
    }
    """

    data = request.get_json()
    app_id = data.get("app_id")
    photo_del_id = data.get("photo_del_id")
    try:
        apps = Applications.query.filter(Applications.app_id == app_id, Applications.app_time_end == None, ).first()
        if apps is None:
            return jsonify(operation_res_build("此申请不存在！", False)), False
        app_ = applications_schema_single.dump(apps)
        flag = del_photo_id(app_, photo_del_id)
        if flag == True:
            return jsonify(operation_res_build("删除成功！", True)), True
        else:
            return jsonify(operation_res_build("图片不存在！", True)), True
    except Exception as e:
        logger.error(request.environ.get('HTTP_X_REAL_IP', request.remote_addr), "a error happened in operation")
        logger.exception(e)
        return jsonify(operation_res_build("a error happened in operation", False)), False


def del_photo_id(app, photo_id):
    model_data = get_model(app["type_name_copy"])
    photos_model = model_data["Photos"]
    photos_schema_single = model_data["Photos_schema_single"]
    photo_key = model_data["photo_key"]
    file_path = model_data["file_path"]
    if photos_model is not None and photo_id is not None:
        # 删除photo
        pre_data = photos_model.query.filter(photo_key == photo_id).first()
        if pre_data is None:
            return False
        photo_data = photos_schema_single.dump(pre_data)
        del_path = photo_data[file_path]
        if os.path.exists(del_path):
            os.remove(del_path)
        db.session.delete(pre_data)
        db.session.commit()
    return True


def delete_receipt(app):
    # 返回所需模型
    model_data = get_model(app["type_name_copy"])
    model_ = model_data["Model"]
    model_schema_single = model_data["Model_schema_single"]
    receipt = model_data["receipt"]
    pre_sql = model_.query.filter(app["app_id"] == model_.app_id)
    if pre_sql.first() is not None:
        item_data_ = model_schema_single.dump(pre_sql.first())

        # photo_key_id = list(item_data_.keys())[0]
        if item_data_[receipt] is not None:
            del_path = item_data_[receipt]
            if os.path.exists(del_path):
                os.remove(del_path)

            pre_sql.update({receipt: None})
            db.session.commit()

    return True
